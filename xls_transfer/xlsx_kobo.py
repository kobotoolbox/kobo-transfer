import re
import string

import openpyxl
from xml.etree import ElementTree as ET

from transfer.media import rename_media_folder
from transfer.xml import generate_new_instance_id


def open_xlsx(excel_file_path):
    """opens xlsx, and returns workbook"""
    try:
        workbook = openpyxl.load_workbook(excel_file_path, read_only=True)
    except FileNotFoundError as e:
        print(f"⚠️ Error: Excel file not found at path '{excel_file_path}'.")
        raise (e)
    except openpyxl.utils.exceptions.InvalidFileException:
        print(f"⚠️ Error: Invalid file format for '{excel_file_path}'")
        raise (e)
    except TypeError as e:
        if "NoneType" in str(e):
            print(f"⚠️ Error: File path is None. Specify xls file path with flag -ef")
            raise (e)
    except Exception as e:
        print(f"⚠️ Something went wrong when reading xlsx file: {e}")
        raise (e)
    return workbook


def create_xml_element_and_tag(parent, tag, text, namespace=None):
    if tag is None:
        raise Exception("Something went wrong.")

    if parent is None and text is None:
        if namespace is None:
            new_element = ET.Element(tag)
        else:
            new_element = ET.Element(
                tag, namespace
            )  # condition only entered for the nsmap dict
    elif parent is None:
        new_element = ET.Element(tag)
        new_element.text = text
    elif text is None:
        new_element = ET.SubElement(parent, tag)
    else:
        new_element = ET.SubElement(parent, tag)
        new_element.text = text

    return new_element


def add_formhub_element(nsmap_dict, formhub_uuid):
    """
    creates and returns initial element for the submission )example of the format is <aDVansdqUpnpUhGxy8 id="aDVansdqUpnpUhGxy8" version="3 (2022-03-08 10:33:24)">)
    and the formhub element nested within it.

    :param nsmap_dict: dictionary containing uid and version
    """
    uid = nsmap_dict["id"]
    submission_xml = create_xml_element_and_tag(None, uid, None, nsmap_dict)
    fhub_element = create_xml_element_and_tag(submission_xml, "formhub", None)
    create_xml_element_and_tag(fhub_element, "uuid", str(formhub_uuid))
    return submission_xml


def add_meta_element(submission_xml, formatted_uuid):
    """creates element in xml with meta tag. meta contains instanceId and deprecatedId elements and appears at end of each submission.
    once meta element is created, it is appended to the xml of whole submission, and then returned.
    method is only called when $edited value is true for the submission

    :param subnmission_xml is the xml representing a single submission. it is built upon and starts with the inital nsmaplement
    :param formatted_uuid is unique instance id of the submission
    """

    # if its #edited, create a new uuid, and  move the old one to the deprecatedID
    # if there is no uuid or uuid is blank, this means it's an initial submission. no deprecated element created, only new uuid (instanceID)
    meta = create_xml_element_and_tag(None, "meta", None)
    instanceId = create_xml_element_and_tag(meta, "instanceID", None)

    if (
        formatted_uuid is None
    ):  # uuid will be generated here when xlsx doesn't contain header _uuid
        formatted_uuid = generate_new_instance_id()[0]
        instanceId.text = formatted_uuid
    else:
        create_xml_element_and_tag(meta, "deprecatedID", str(formatted_uuid))
        instanceId.text = str(generate_new_instance_id()[1])
        formatted_uuid = instanceId.text

    submission_xml.append(meta)

    return formatted_uuid


def initialize_elements():
    """
    creates and returns initial parent xml elements for data.
    root and results will be appended to as xlsx is parsed.
    """
    root = create_xml_element_and_tag(None, "root", None)
    results = create_xml_element_and_tag(None, "results", None)
    return root, results


def get_question_headers(headers):
    """filters the column labels/headers, and returns only the ones that contain questions"""
    question_headers = []
    added_on_headers_during_export = [
        "_id",
        "_uuid",
        "_submission_time",
        "_validation_status",
        "_notes",
        "_status",
        "__version__",
        "_submitted_by",
        "_tags",
        "_index",
        "$edited",
        "_parent_table_name",
        "_parent_index",
        "_submission__id",
        "_submission__uuid",
        "_submission__submission_time",
        "_submission__validation_status",
        "_submission__notes",
        "_submission__status",
        "_submission__submitted_by",
        "_submission___version__",
        "_submission__tags",
    ]
    recent_question = None
    for header in headers:
        geopoint = is_geopoint_header(str(recent_question), header)
        if geopoint or header in added_on_headers_during_export:
            continue
        recent_question = header
        question_headers.append(header)
    return question_headers


def is_geopoint_header(recent_question, col_name):
    """
    returns false if column header should be treated should be treated as a question and
    returns true if column header should be ignored since it is part of geopoint/geoline type
    used to filter out headers that follow pattern of _<question_name>_latitude etc.
    """
    geopoint_patterns = (
        r"_" + re.escape(recent_question) + r"_(latitude|longitude|altitude|precision)"
    )
    return re.match(geopoint_patterns, col_name) is not None


def create_group(group_names_arr, cell_value, parent_group=None):
    """
    Creates and returns an xml element of nested groups.

    :param group_name Header containing group split by ('/'). Example of xlsx header that would be split is [pets/pet/name_of_pet]
    :param cell_value: value stored in the cell for a particular submission, under the group_name column.
    :param parent_group: None when the first string in group_name is the root.
    """
    if parent_group is None:
        initial = create_xml_element_and_tag(None, group_names_arr[0], None)
        parent_group = initial
    else:
        initial = parent_group

    group_element = None
    for group in group_names_arr:
        if parent_group.tag == group:
            continue
        group_element = parent_group.find(".//" + group)

        if group_element == None:
            group_element = create_xml_element_and_tag(parent_group, group, None)
        if (
            group == group_names_arr[-1]
        ) and cell_value:  # last element in group_name is the question
            group_element.text = str(cell_value)

        parent_group = group_element
        group_element = None

    return initial


def extract_uuid(cell_value):
    """
    if uuid is not present in xlsx, it generates a uuid for the submission
    othrerwise, it extracts uuid from cell and returns it
    """
    formatted_uuid = None
    if cell_value:
        formatted_uuid = "uuid:" + str(cell_value)
    return formatted_uuid


def extract_submission_data(submission_data):
    formhub_uuid = submission_data["formhub_uuid"]
    __version__ = submission_data["__version__"]
    return formhub_uuid, __version__


def create_nsmap_dict(submission_data):
    """extracts version and asset uid from submission_data to create
    dictionary containing them. this dictionary (nsmap_dict) will be used to create the first parent xml element
    for a single submission"""
    v = submission_data["version"]
    uid = submission_data["asset_uid"]
    nsmap_dict = {
        "id": str(uid),
        "version": str(v),
    }
    return nsmap_dict


def process_data_in_columns(submission_xml, col_name, cell_value):
    """
    creates the xml for a single submission which is represented in a single row in the xlsx

    :param submission_xml is the root element for the single submission that will be appended to as rows are parsed
    :param col_name is the header/label of the column in xlsx
    :param cell_value is the cell value in xls for current row/submission, in col_name

    returns:
    all_empty is a boolean that is true when the submission is blank and all questions have not been responded to
    """
    all_empty = None
    if cell_value in [None, "None", "none"]:
        cell_value = ""
    else:
        all_empty = False
    if (
        "/" in col_name
    ):  # column headers that include / indicate {group_name}/{question}
        submission_xml = create_group(
            col_name.split("/"), str(cell_value), submission_xml
        )
    else:
        if col_name in ["end", "start"]:
            cell_value = cell_value.isoformat() if cell_value else ""
        create_xml_element_and_tag(submission_xml, col_name, str(cell_value))

    return all_empty


def process_single_row(row, headers, submission_xml):
    """handles a single submission (represented in one row of xls)
    iterates through cells in the row and create corresponding XML elements

    returns:
    _index value of the submission
    the uuid (if one exists)
    all_empty boolean value which is true when submission is empty/blank"""
    all_empty = True
    index = str(row[headers.index("_index")])
    question_headers = get_question_headers(headers)
    formatted_uuid = None
    if "_uuid" in headers:
        formatted_uuid = extract_uuid(str(row[headers.index("_uuid")]))

    for col_num, cell_value in enumerate(row, start=1):
        col_name = headers[col_num - 1]
        if col_name in question_headers:
            all_empty = process_data_in_columns(submission_xml, col_name, cell_value)
            if (
                not all_empty
            ):  # all_empty should only be true when all cell values are blank
                all_empty = False

    return index, formatted_uuid, all_empty


def add_version_and_meta_element(submission_xml, formatted_uuid, __version__):
    """
    creates xml elements at the end of submission stating version number and meta data
    """
    version_element = create_xml_element_and_tag(None, "__version__", __version__)
    submission_xml.append(version_element)
    formatted_uuid = add_meta_element(submission_xml, formatted_uuid)
    return formatted_uuid


def add_prev_next(root):
    """
    creates xml elements that appear prior to the results data (count, next, previous) and appends it
    """
    create_xml_element_and_tag(root, "next", None)
    create_xml_element_and_tag(root, "previous", None)
    return root


def general_xls_to_xml(excel_file_path, submission_data, warnings=False):
    """
    parses entire xlsx and creates xml from it
    returns: root of the xml that can be used to transfer to kobo project
    """
    workbook = open_xlsx(excel_file_path)
    # first sheet should have all data, if xlsx contains other sheets, they must be for repeat groups
    sheet = workbook.worksheets[0]

    formhub_uuid, __version__ = extract_submission_data(submission_data)
    nsmap_dict = create_nsmap_dict(submission_data)
    root, results = initialize_elements()
    headers = [cell.value for cell in sheet[1]]

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        submission_xml = add_formhub_element(nsmap_dict, formhub_uuid)
        # error will be raised if no $edited column is found
        if not eval(  # all false or blank $edited submissions are being skipped
            str(row[headers.index("$edited")])
        ):
            continue

        index, formatted_uuid, all_empty = process_single_row(
            row, headers, submission_xml
        )
        if all_empty:
            print(
                "Warning: Data may include one or more blank responses where no questions were answered."
            )
        submission_xml = xml_from_repeat_sheets(
            submission_xml, workbook, index
        )  # iterate through other sheets to create repeat groups, and append to xml
        formatted_uuid = add_version_and_meta_element(
            submission_xml, formatted_uuid, __version__
        )

        # for initial transfer (without uuids), each submission associated with index
        # index folder is renamed to uuid to complete transfer to kobo and associate attachment to specific response
        rename_media_folder(submission_data, formatted_uuid, index)
        results.append(submission_xml)

    root = add_prev_next(root)
    root.append(results)
    # test_by_writing(root)
    workbook.close()
    return root


def test_by_writing(root):
    root = ET.ElementTree(root)
    root.write("./submission.xml")


def find_nth_tag(xml, n, element_tag):
    """
    finds nth occurence of tag in xml and returns the element
    used to append data to correct repeat group
    """
    occurrences = 0
    for element in xml.iter(element_tag):
        occurrences += 1
        if occurrences == int(n):
            return element
    return None


def get_sheet_info(headers):
    """
    method is only called when there are repeat groups
    extracts and returns the index value of _index, _parent_index, and _parent_table_name, within
    the headers list"""
    try:
        index_header = headers.index("_index")
        parent_index_header = headers.index("_parent_index")
        parent_table_header = headers.index("_parent_table_name")
    except Exception:
        print(
            "Error: if xlsx file has multiple tabs, data in extra sheets must be in expected repeating group format"
        )
        raise
    return index_header, parent_index_header, parent_table_header


def add_groups_if_missing(question_headers, sheet_names, submission_xml):
    """when repeat groups are nested within groups, and groups haven't been created yet (because they don't appear in
    first xls sheet), they need to be created an added to xml before proceeding with repeat groups.
    this method checks if all groups that have repeat groups nested within them are part of xml, and creates them if they aren't
    """
    non_repeat_groups = []
    for column_label in question_headers:
        group_names_arr = column_label.split("/")
        for index, group in enumerate(group_names_arr):
            if str(group) in sheet_names:  # sheet names are the repeat group names
                # if groups exist before first repeat group in column label [group1/group2/repeatgroup/..], append to list
                non_repeat_groups.append(group_names_arr[:index])
                break
    # create non-repeating group in submission_xml if it doesn't exist yet
    for group in non_repeat_groups:
        create_group(group, None, submission_xml)


def create_repeat_group_xml_element(current_sheet_name, question_headers, headers, row):
    """
    creates the xml element for a single row in the repeat group xls sheet

    returns:
    repeat_sheet_xml_element is the xml for the current repeat sheet. each sheet in xls name should be the same as repeat group.
    when column label might be group1/group2/repeatgroup/repeatgroup2/question1, and the sheet name is repeatgroup2, xml element repeatgroup2 (with question1 nested) will be returend
    index_of_sheet_group is the index where sheet name is found in column label. in above example, when column label is split('/'), index is 3
    parent_of_sheet_group is the group preceding sheet name repeat group. in above example, this is repeatgroup

    """
    repeat_sheet_xml_element = None
    index_of_sheet_group = None
    parent_of_sheet_group = None
    for col_name in question_headers:
        cell_value = str(row[headers.index(col_name)])

        if cell_value in [None, "None", "none"]:
            cell_value = ""

        group_names = col_name.split("/")
        # split into array of header, starts creating repeat xml element from when repeat sheet name is mentioned
        index_of_sheet_group = group_names.index(str(current_sheet_name))

        if (
            repeat_sheet_xml_element is None
        ):  # this is for first column of the sheet, creates all elements in header starting from spreadsheet name
            repeat_sheet_xml_element = create_group(
                group_names[index_of_sheet_group:], str(cell_value)
            )
        else:  # following columns of the sheet (nest the elements, or append to the elements created from first column)
            repeat_sheet_xml_element = create_group(
                group_names[index_of_sheet_group:],
                str(cell_value),
                repeat_sheet_xml_element,
            )
    if index_of_sheet_group != 0:
        parent_of_sheet_group = group_names[index_of_sheet_group - 1]

    return repeat_sheet_xml_element, index_of_sheet_group, parent_of_sheet_group


def xml_from_repeat_sheets(submission_xml, workbook, submission_index):
    """method is called when there are multiple sheets in xlsx, because it is assumed to be repeat groups
    method iterates through the xls sheets and adds each repeat group element to the xml
    limitation: sheet needs to be ordered from parent —> child; the parent_table must precede child in xls sheet order
    """
    sheet_names = workbook.sheetnames
    parent_indexes = []  # cleared every time its a new sheet.
    for sheet_name in sheet_names[1:]:
        new_indexes = []
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]

        index_header, parent_index_header, parent_table_header = get_sheet_info(headers)
        question_headers = get_question_headers(headers)
        add_groups_if_missing(
            question_headers, sheet_names, submission_xml
        )  # ensure all non-repeating group headers exist in submission_xml

        for row in sheet.iter_rows(min_row=2, values_only=True):
            index = str(row[index_header])
            parent_index = str(row[parent_index_header])  # row's parent value
            parent_table = str(row[parent_table_header])
            parent_is_first_sheet = False

            if parent_table == str(sheet_names[0]):
                parent_is_first_sheet = True

            if parent_is_first_sheet:
                if parent_index != submission_index:
                    continue
                # populate parent_indexes with index of the rows that have submission_index passed in from first sheet
                parent_indexes.append(index)

            else:
                # submission row only relevant if its part of the parent_indexes
                if parent_index not in parent_indexes:
                    continue
                new_indexes.append(
                    index
                )  # mantain relevant indexes for this submission for next sheet (where current sheet might be parent)

            repeat_sheet_xml_element, index_of_sheet_group, parent_of_sheet_group = (
                create_repeat_group_xml_element(
                    sheet_name, question_headers, headers, row
                )
            )

            if question_headers != []:
                element_to_append_to = None
                if (
                    parent_is_first_sheet and index_of_sheet_group == 0
                ):  # repeat group is not nested
                    element_to_append_to = submission_xml
                elif parent_is_first_sheet and index_of_sheet_group != 0:
                    element_to_append_to = submission_xml.find(
                        ".//" + str(parent_of_sheet_group)
                    )
                else:
                    element_to_append_to = find_nth_tag(
                        submission_xml,
                        parent_indexes.index(parent_index) + 1,
                        parent_of_sheet_group,
                    )
                element_to_append_to.append(repeat_sheet_xml_element)

                if row == sheet.max_row:
                    parent_indexes = new_indexes

    return submission_xml
