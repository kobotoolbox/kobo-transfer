import string
import re
from datetime import datetime
from xml.etree import ElementTree as ET
import openpyxl
import pandas as pd
from dateutil import parser
import requests
from helpers.config import Config
from transfer.xml import get_src_submissions_xml
from .media import rename_media_folder
from .xml import generate_new_instance_id


def open_xlsx(excel_file_path):
    """opens xlsx, and returns workbook"""
    try:
        workbook = openpyxl.load_workbook(excel_file_path, read_only=True)
    except FileNotFoundError as e:
        print(f"⚠️ Error: Excel file not found at path '{excel_file_path}'.")
        raise (e)
    except openpyxl.utils.exceptions.InvalidFileException:
        print(f"⚠️ Error: Invalid file format for '{excel_file_path}'")
        raise (e)
    except TypeError as e:
        if "NoneType" in str(e):
            print(f"⚠️ Error: File path is None. Specify xls file path with flag -ef")
            raise (e)
    except Exception as e:
        print(f"⚠️ Something went wrong when reading xlsx file: {e}")
        raise (e)
    return workbook


def add_formhub_element(nsmap_element, formhub_uuid):
    """creates formhub element with nested uuid"""
    uid = nsmap_element["id"]
    _uid = ET.Element(uid, nsmap_element)
    fhub_el = ET.SubElement(_uid, "formhub")
    uuid_el = ET.SubElement(fhub_el, "uuid")
    uuid_el.text = formhub_uuid
    return _uid


def add_meta_element(_uid, formatted_uuid):
    """meta tag
          <meta>
                <instanceID>uuid:a0ea37ef-ac71-434b-93b6-1713ef4c367f</instanceID>
                <deprecatedID>
            </meta>
    }"""
    meta = ET.Element("meta")
    if formatted_uuid is None: #uuid will be generated here when xlsx doesn't contain header _uuid
        #formatted_uuid = "uuid:"
        formatted_uuid = generate_new_instance_id()[0]
    instanceId = ET.SubElement(meta, "instanceID")
    deprecatedId = ET.SubElement(meta, "deprecatedID")
    instanceId.text = str(formatted_uuid)
    deprecatedId.text = str(formatted_uuid)
    _uid.append(meta)
    return formatted_uuid


def extract_uuid(cell_value):
    formatted_uuid = None
    if not cell_value:  # if there is no uuid specified, new one will be generated
        formatted_uuid = generate_new_instance_id()[0] 
    else:
        formatted_uuid = 'uuid:' + str(cell_value) #TODO CHANGED HERE
    return formatted_uuid


def single_submission_xml( _uid, col_name, cell_value, all_empty
):
#TODO, OTHER THAN THE IF CONDITION, FORMATTED_UUID ALWAYS RETURNS NONE
    formatted_uuid = None
    if cell_value in [None, 'None', 'none']:
        cell_value = ""
    else:
        all_empty = False
    
    # if xlsx data is downloaded from kobo, it will contain this column
    if col_name == "_uuid":
        formatted_uuid = extract_uuid(cell_value)
        #return all_empty, formatted_uuid
    
    # column headers that include / indicate {group_name}/{question}
    elif len(col_name.split('/')) >= 2: 
        _uid = combine_attempt(col_name.split('/'), str(cell_value), _uid)
    else: 
        cell_element = ET.SubElement(_uid, col_name)
        if col_name in ['end', 'start']:
            #if cell_value:
                cell_value = cell_value.isoformat() if cell_value else ""
        cell_element.text = str(cell_value)

    return all_empty, formatted_uuid


"""
THIS IS THE OLD ONE

def single_submission_xml( _uid, col_name, cell_value, all_empty
):
#TODO, OTHER THAN THE IF CONDITION, FORMATTED_UUID ALWAYS RETURNS NONE
    if cell_value in [None, 'None', 'none']:
        cell_value = ""
    else:
        all_empty = False
    
    # if xlsx data is downloaded from kobo, it will contain this column
    if col_name == "_uuid":
        formatted_uuid = extract_uuid(cell_value)
        return all_empty, formatted_uuid

    #formatted_uuid = None
    #if col_name == "_uuid":
        #if not cell_value:  # if there is no uuid specified, new one will be generated
         #   formatted_uuid = generate_new_instance_id()[0] #TODO CHANGED HERE
       # else:
        #    formatted_uuid = 'uuid:' + str(cell_value) #TODO CHANGED HERE
        #return all_empty, formatted_uuid

    # column headers that include / indicate {group_name}/{question}
    group_arr = col_name.split("/")
    if len(group_arr) >= 2: 
        _uid = combine_attempt(group_arr, str(cell_value), _uid)
        return all_empty, formatted_uuid
    
    cell_element = ET.SubElement(_uid, col_name)
    if col_name in ['end', 'start']:
        if cell_value:
            cell_value = cell_value.isoformat()
    cell_element.text = str(cell_value)
    return all_empty, formatted_uuid"""



def is_geopoint_header(recent_question, col_name):
    geopoint_patterns = r"_" + re.escape(recent_question) + r"_(latitude|longitude|altitude|precision)"
    return re.match(geopoint_patterns, col_name) is not None


def extract_submission_data(submission_data):
    formhub_uuid = submission_data["formhub_uuid"]
    __version__ = submission_data["__version__"]
    return formhub_uuid, __version__


def create_nsmap_element(submission_data):
    v = submission_data["version"]
    uid = submission_data["asset_uid"]
    nsmap_element = {
        "id": str(uid),
        "version": str(v),
    }
    return nsmap_element

# Iterate through cells in the row and create corresponding XML elements
def process_single_row(row, headers, added_on_headers_during_export, _uid):
    all_empty = True
    index = None
    recent_question = None
    for col_num, cell_value in enumerate(row, start=1):
        col_name = headers[col_num - 1]
        if col_name == '_index': 
            index = str(cell_value)
        if not col_name:
            continue
        geopoint = is_geopoint_header(str(recent_question), col_name)
        if geopoint or col_name in added_on_headers_during_export: 
            continue
        recent_question = col_name
        all_empty, formatted_uuid = single_submission_xml(
            _uid, col_name, cell_value, all_empty
        )
    return index, formatted_uuid, all_empty

def initialize_elements():
    #added and works
    root = ET.Element("root")
    results = ET.Element("results")
    return root, results

def t_general_xls_to_xml(
    excel_file_path, submission_data, warnings=False
):
    workbook = open_xlsx(excel_file_path)
    sheet = workbook.worksheets[
        0
    ]  # first sheet should have all data, if xlsx contains other sheets, they must be for repeat groups

    formhub_uuid, __version__= extract_submission_data(submission_data)
    nsmap_element = create_nsmap_element(submission_data)
    root, results = initialize_elements()
    headers = [cell.value for cell in sheet[1]]
    # columns automatically generated with kobo (this is after data has been downloaded from kobo)
    added_on_headers_during_export = ['_id', '_submission_time', '_validation_status', '_notes',	'_status',	'__version__', '_submitted_by', '_tags', '_index']

    if warnings:
        kobo_xls_match_warnings(headers, submission_data)
    
    num_results = 0
    for row_num, row in enumerate(
        sheet.iter_rows(
            min_row=2,
            values_only=True,
        ),
        start=2,
    ):
        _uid = add_formhub_element(nsmap_element, formhub_uuid)
        #all_empty = True
        #index = None
        #recent_question = None
        index, formatted_uuid, all_empty = process_single_row(row, headers, added_on_headers_during_export, _uid)        
        if all_empty:
            print(
                "Warning: Data may include one or more blank responses where no questions were answered."
            )

        # iterate through other sheets to create repeat groups, and append to xml
        repeat_elements = new_repeat(_uid, workbook, index)
        if repeat_elements is not None:
            print('here')
            _uid = repeat_elements

        formatted_uuid = add_version_and_meta_element(_uid, formatted_uuid, __version__)

        # for initial transfer (without uuids), each submission associated with index
        # index folder is renamed to uuid to complete transfer to kobo and associate attachment to specific response
        rename_media_folder(submission_data, formatted_uuid, index) 
        results.append(_uid)
        num_results += 1

    root =  add_initial_elements(root, num_results, results)
    test_by_writing(root)
    workbook.close()
    return root

def test_by_writing(root):
    root = ET.ElementTree(root)
    root.write("./um.xml")

def add_version_and_meta_element(_uid, formatted_uuid, __version__):
    version = ET.Element("__version__")
    version.text = __version__
    _uid.append(version)
    formatted_uuid = add_meta_element(_uid, formatted_uuid)
    return formatted_uuid


def add_initial_elements(root, num_results, results):
    count = ET.SubElement(root, "count")
    count.text = str(num_results)
    next = ET.SubElement(root, "next")
    next.text = None
    previous = ET.SubElement(root, "previous")
    previous.text = None
    root.append(results)
    return root

#TODO ur combined attempt is causing problems
def combine_attempt(group_name, cell_value, parent_group = None):
    if parent_group is None:
        initial = ET.Element(group_name[0])
        #group_name = group_name[1:] #root is initial, search for elements after it when you use .find//
        #alternatively you can do the cnotinue thing... that you did in nested
        parent_group = initial
        print("not nested1")
    else:
        initial = parent_group
        print("nested1")
    
    group_element = None
    for group in group_name: 
        if parent_group.tag == group:
            continue
    
        group_element = parent_group.find(".//" + group)

        if (group_element == None):
            group_element = ET.SubElement(parent_group, group)

        if (group == group_name[-1]): #last element in group_name is the question
            group_element.text = str(cell_value)
        
        parent_group = group_element
        group_element = None
 
    return initial
    


#TODO YOU ATTEMPTED TO OMBINE THEM ABOVE BUT THESE METHODS CURRENTLY WORK
"""def nested_group_element(_uid, group_name, cell_value):
    print("nested)")
    parent_group = _uid
    #if (_uid == None):
     #   parent_group = ET.Element(group_name[0]) 
      #  group_name = group_name[1:]
    group_element = None
    for group in group_name: 
        if parent_group.tag == group:
            continue
    
        group_element = _uid.find(".//" + group)

        if (group_element == None):
            group_element = ET.SubElement(parent_group, group)

        if (group == group_name[-1]): #last element in group_name is the question
            group_element.text = str(cell_value)
        
        parent_group = group_element
        group_element = None
    
    return _uid

def group_section(group_name, cell_value):
    print("not nested)")
    #group_name = group.split("/")
    initial = ET.Element(group_name[0]) 
    #group_name = group_name[1:] #root is initial, search for elements after it when you use .find//
    #alternatively you can do the cnotinue thing... that you did in nested
    
    parent_group = initial
    group_element = None
    for group in group_name: 

        if parent_group.tag == group: #TODO remember that you removed group_name = group_name[1:] and replaced it with this... IF YOU GET AN ERROR FOR ONE OF THE REPEAT STUFF THIS MIGHT BE WHY
            continue 

        group_element = parent_group.find(".//" + group) 
        if (group_element == None):
            group_element = ET.SubElement(parent_group, group)
        
        if (group == group_name[-1]): #last element in group_name is the question
            group_element.text = str(cell_value)
        
        parent_group = group_element
        group_element = None

    return initial"""

def find_n(xml, n, element_name):
    occurrences = 0
    for element in xml.iter(element_name):
        occurrences += 1
        if occurrences == int(n):
            return element
    return None

def check_for_group(_uid, group_name):
    parent_group = _uid
    group_element = None
    for group in group_name: 

        group_element = _uid.find(".//" + group)

        if (group_element == None):
            group_element = ET.SubElement(parent_group, group)
        
        parent_group = group_element
        group_element = None
    
    return _uid

"""def create_repeat_element(submission_xml, group_names):
    parent_group = _uid
    #if (_uid == None):
     #   parent_group = ET.Element(group_name[0]) 
      #  group_name = group_name[1:]
    group_element = None
    for group in group_name: 
        if parent_group.tag == group:
            continue
    
        group_element = _uid.find(".//" + group)

        if (group_element == None):
            group_element = ET.SubElement(parent_group, group)

        if (group == group_name[-1]): #last element in group_name is the question
            group_element.text = str(cell_value)
        
        parent_group = group_element
        group_element = None
    
    return _uid
"""

def new_repeat(submission_xml, workbook, submission_index):
        """method is called when there are multiple sheets in xlsx, because it is assumed to be repeat groups"""
        #uuid = uuid[len("uuid:") :]
        sheet_names = workbook.sheetnames 
        #sheet_names = sheet_names[1:]

        original_indexes = [] #cleared every time its a new sheet. 

        for sheet_name in sheet_names[1:]:
        #for sheet_name in sheet_names:
            new_indexes = []
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in sheet[1]] 
            try:
               # submission_uid_header = headers.index("_submission__uuid")
                index_header = headers.index("_index")
                parent_index_header = headers.index('_parent_index')
                parent_table_header = headers.index('_parent_table_name')
            except Exception:
                print(
                    "Error: if xlsx file has multiple tabs, data in extra sheets must be in expected repeating group format"
                )
                raise

            question_headers = [item for item in headers if not item.startswith('_')] 

            #assumes question headers all have same path, but this is not true
            #if there is group within a repeat group, question headers might be different

            #[aw4/pet, aw4/ve/q]

            #loop through the question headers and make sure that none of them start with a non-repeating group
            #if they do start with a non-repeating group, make sure that its in the submission_uid already
            #if its not, you need to create it, and cut the header to start from the repeating group !
            #list1 = ['oh/my/lol', 'oh/no/my/lol', 'oh/what/my/okay/lol']
            #list2 = ['my', 'lol']
            
            non_repeat_groups = []
            
            for column in question_headers:
                parts = column.split('/')
                for part in parts:
                    if part in sheet_names:
                        break
                    else:
                        non_repeat_groups.append(part)
                        break

            
            for normal_group in non_repeat_groups:
                check_for_group(submission_xml, normal_group.split('/'))
        

            #print(question_headers)
           

            #for case3 there are no question headers
            #you can't create the element because you don't know where they are going to go... 
            #if they are nested you have nowhere to put them. 



            #you don't want to create any elements
            #you want to make sure that you are looking at relevant indexes tho
            #check if the parent table is equal to the name of the first sheet
            #if it is, look for submissionid match
            #save these indexes as original indexes
            #continue to next sheet 



            #print(non_repeat_groups)
            #for each of the elements in non_repeat_groups, make sure it is in submission_uid
            #if it is, do nothing
            #if it is not, edit submission_uid and add the group
            #create the element in submission_uid if it does not already exist.. 
            for row in sheet.iter_rows(min_row=2, values_only=True):
                mini_group = None
                #submission_uid = str(row[submission_uid_header])
                index = str(row[index_header]) #this is the current value
                parent_index = str(row[parent_index_header]) #this is row's parent value
                parent_table = str(row[parent_table_header])
                group_to_appendto = None



                #so you broke it up into one condition where its the og
                #so parents have to match passed in

                #other condition is when parent refers to a different sheet, not og sheet


                #if starting element of the group is the repeat group (same as sheet name)
                #this indicates that it's parent will be submission_uid
                #it will be directly appended, not nested in anything

                #if question_headers == []:
                     #for case3 there are no question headers
                    #you don't want to create any elements
                    #you want to make sure that you are looking at relevant indexes tho
                    #check if the parent table is equal to the name of the first sheet
                    #if it is, look for submissionid match
                    #save these indexes as original indexes
                    #continue to next sheet
                    #if parent_table == str(sheet_names[0]):
                        #if parent_index != submission_index:
                         #   continue
                        #else: 
                         #   original_indexes.append(index)
                    #continue

             #   print(parent_table)
              #  print(str(sheet_names[0]))
                
                #if str(sheet_name) == str(question_headers[0].split('/')[0]):  #because the string ur getting is home..., however no sheet_name called home
                if parent_table == str(sheet_names[0]):
                    #only look at rows with same value as this submission (passed in as parameter)
                    if parent_index != submission_index: 
                        continue
                    else:
                
                        #populate original_indexes with index of the rows that haave parent_group == passed in value from first sheet

                        original_indexes.append(index) #first sheet, makes sure this is not empty.. 
                        
                        #for case3 there are no question headers
                        #you don't want to create any elements
                        #you want to make sure that you are looking at relevant indexes tho
                        #check if the parent table is equal to the name of the first sheet
                        #if it is, look for submissionid match
                        #save these indexes as original indexes
                        #continue to next sheet
                
                #currently this condition is whenever repeating group is not being directly appended to the submission_xml
                #the repeat group for this sheet is nested
                #for josh's case, it's not being directly appended to submission_xml, because parent is 'home'
                #TODO: need to add a condition here
                #element_nonrepeat_group = None

                #nested, but if current repeat group is not a repeat group

                #elif original_indexes == []: #indicates that parent group is not a repeat group
                 #   lookingfor = str(question_headers[0].split('/')[0])
                  #  group_to_appendto = submission_xml.find(".//" + lookingfor)
    
                    #therefore, the parent must be a group in the submission_xml
                    #group_tolook_for = question_headers[0].split('/')[0]
                    #print("yas")
                

                #maybe

                #repeat group is nested, therefore parent indexes will be another repeat group
                #if its another sheet, only relevant columns should be the ones where parent_index is in original_indexes list
                else: 
                    #TODO: josh case comes here so thats good
                    #in josh case original_indexes is empty... therefore never continues
                    #in josh case it NEVER goes into the first condition because its doing the if condition based on sheet_name
                    #need to populate original indexes.... 


                    print(original_indexes)
                    if parent_index not in original_indexes:  #you have to see if its part of the original indexes
                        continue
                    else: 
                        new_indexes.append(index)

        
                #iterating through each column
                #so each mini group only refers to a single row
                #its refreshed to None at every row iter
                #so group op5 is created for the first column with group section
                #and then groupwm is created
                #and then you are now looking at wm95, and you have to append it to the group op5 etc..

                #if question_headers == []:
                    #create the group 
                    #but you can't create the gorup... 

                #now ur in the next sheet 

                
                for col_name in question_headers: 
                    col_num = headers.index(col_name)
                    cell_value = str(row[col_num])

                    if cell_value in [None, 'None', 'none']:
                        cell_value = ""

                    group_names = col_name.split('/') 

                    #now there is an array of the header
                    #you start the repeat stuff from when repeat sheet name is mentioned!
                    index_of_sheet_group = group_names.index(str(sheet_name)) 
                    
                    """
                    JOSH CASE WORKED WITHOUT THIS??? SO FIGURE OUT IF YOU EVEN NEED IT
                    BOTH W/O AND WITH GAVE SAME RESUL
                    if index_of_sheet_group != 0:
                        parent_of_sheet_group = group_names[index_of_sheet_group - 1]
                        #when a repeat group is within a group... 
                        if parent_of_sheet_group not in sheet_names: #when parent group is not a repeat group
                            group_to_appendto = submission_xml.find(".//" + str(parent_of_sheet_group))
                    """
                    #TODO 
                    #all code below is dependent on sheet name since mini group starts from wtvr the sheetname/repeating group is
                    #when there is group within a repeat group, this is ok, since can still start the elements from the sheetname
                    #HOWEVER, need different condition when the repeat group is nested in a group
                    if mini_group is None: #this is for first column of the sheet, creates all elements in header starting from spreadsheet name
                        mini_group = combine_attempt(group_names[index_of_sheet_group:], str(cell_value)) #this is from the top
                        #mini_group = group_section(group_names[index_of_sheet_group:], str(cell_value)) #this is from the top
                    else: #following columns of the sheet (nest the elements, or append to the elements created from first column)
                        mini_group = combine_attempt(group_names[index_of_sheet_group:], str(cell_value), mini_group)
                        #mini_group = nested_group_element(mini_group, group_names[index_of_sheet_group:], str(cell_value)) #TODO OOP THIS DOESNT SEEM TO BE THE PROBLEM
                

                    #test_by_writing(mini_group)

                #group_names is going to be the last column of the current sheet
                #group_names[0] would be the first parent group of the last q in the sheet
                #i think this is done w the assumption that all in a sheet are going to start with the same group?? IDK THO
                if question_headers != []:
                    print(question_headers)
                    if (str(sheet_name) == group_names[0]): #when initial parent repeat group (the one it starts with) is same as sheet. 
                        submission_xml.append(mini_group)
                
                        
                        #so in nestedjosh case2, the header says home/repeat/skdlajf, so it won't enter this. actually that could be ok 
                        #itll enter the next condition
                        #in the next condition, it splits the header... 
                        #finds the occurence of 

                    if (str(sheet_name) != group_names[0]): #if its not the first sheet/first parent element
                        print(sheet_name)
            
                        um = question_headers[0].split('/')
                            #with an extra groups in some columns, above code is fine. 
                            #this is where problems might start.
                            #TODO NOT ALL HEADERS IN SINGLE REPEAT GROUP ARE UNDER SAME PATH.... but it still works lol

                        #TODO: not needed for josh case
                        #if group_to_appendto != None:
                            #element = group_to_appendto
                        


                


                        


                       
                            #ok so when you do um[index_of_sheet_group-1], you are assuming that the parent group is a repeating group
                            #you are finding the one right before your current sheet index (like where in the header ur current sheet is found)
                            #pass in submission to append to, the nth mention of parent to find, parent group of the element created 
                        element = find_n(submission_xml, original_indexes.index(parent_index) + 1, um[index_of_sheet_group-1])

                            #IF CONDITION
                        #WHAT IF REPEAT GROUP HASN'T BEEN CREATED YET..
                        #take all the elements before it 
                        #what if its mutliple repeat groups that havne't been created yet
                        #if parent has not been created yet 
                        #then this wil be blank uh oh.... 
                            #if element is None: 
                             #   index_of_parent_group = group_names.index(index_of_sheet_group-1) 
                              #  element = create_repeat_element(submission_xml, group_names[0:index_of_parent_group])
                              #if the parent hasn't been created yet
                        #we KNOW that it is a repeat group that hasn't been created
                        #check the sheet names... if it is within the sheetname. and it is the parent_table
                        #WHAT IF THERE ARE TWO BLANKISH REPEATGROUPS IN A ROW
                        #TODOTODOTODODOTODO 
                        #TODO MAJOR HERE
                        #so we actually need to append the mini group... 
                        #create the repeat group before it and nest the whole mini group within it
                        #

        
                        element.append(mini_group)

                    if row == sheet.max_row:
                        original_indexes = new_indexes
                
        
       # test_by_writing(submission_xml)
                    

        return submission_xml




"""def repeat_sheet(sheet, headers, target_uuid, current_uuid, question_headers): #this is function i made for the firstr sheet lol
    xml_dict = {}

    parent_header = headers.index('_parent_index')
    parent_index = None
    #parent_index = sheet.cell(row = 2, column = parent_header).value #this is the parent index of the first 

    for row in sheet.iter_rows(min_row=2, values_only=True):
        group_section = None

        if target_uuid != current_uuid: 
            continue
        
        if parent_index == None: #this should only be the first time
            parent_index = row[parent_header]
        
        #parent_index will always be the same for the first sheet

        for question in question_headers:
            col_num = headers.index(question)
            cell_value = str(row[col_num])

            if group_section == None: 
                group_section = group_section(question, cell_value) #xml from the top <op5>
            else: #group section will be no longer be none, when uve created the op5 element already. this is for second question in question header
                group_section = combine_attempt(question, cell_value, group_section)
        
        #done with a single row, so you want to put the index in it yk. 
        xml_dict[parent_index] == group_section
    
    return xml_dict 




def find_last_common_element(root1, root2):
    common_element = None
    
    # Get iterators for both XML trees
    iter1 = root1.iter()
    iter2 = root2.iter()
    
    # Iterate through both trees simultaneously
    for elem1, elem2 in zip(iter1, iter2):
        if elem1.tag == elem2.tag:
            common_element = elem1
        else:
            # Elements diverged, break the loop
            break
    
    return common_element"""



"""

def repeat_try(submission_xml, uuid, workbook):
    method is called when there are multiple sheets in xlsx, because it is assumed to be repeat groups
    uuid = uuid[len("uuid:") :]
    sheet_names = workbook.sheetnames 
    sheet_names = sheet_names[1:]

    xml_dict = {}
    
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]] 
        
        try:
            submission_uid_header = headers.index("_submission__uuid")
            index_header = headers.index("_index")
            parent_index_header = headers.index('_parent_index')
        except Exception:
            print(
                "Error: if xlsx file has multiple tabs, data in extra sheets must be in expected repeating group format"
            )
            raise

        question_headers = [item for item in headers if not item.startswith('_')] #question headers should be the same for a single sheet
#       print(question_headers)

        if sheet_name == sheet_names[0]: #if its the first sheet, call another method
            xml_dict = repeat_sheet(sheet, headers, uuid, submission_uid, question_headers)
            continue
            #you have xml_dict with the indexes that you want...

        initial_parent_index = None #lets say we are dealing with 1 first. 
        same_parent = []

        #now you are starting at the second sheet
        for row in sheet.iter_rows(min_row=2, values_only=True):
            submission_uid = str(row[submission_uid_header])
            index = str(row[index_header])
            parent_index = str(row[parent_index_header]) #this is the current value

            if (parent_index not in xml_dict.keys()):
                continue

            if initial_parent_index == None: #this should only be the first time (only happens once)
                initial_parent_index = parent_index 

            xml_group = xml_dict[initial_parent_index] #get the relevant xml_group. this is the 1 index group.. 
            group = None
            if initial_parent_index == parent_index: #this means we are dealing w same xml group, just creating sections, and appending to it, lets say this is 1 for the first row. 
                #this row's parent index is 1
                for col_name in question_headers: 
                    col_num = headers.index(col_name)
                    cell_value = str(row[col_num])
                # "group_op5ep98/group_wm95d90/group_sw2vc60/nested_x_2_within_re_p_this_also_repeats"
                    if group == None: 
                        group = group_section(col_name, cell_value) #this should create it from the top
                    else: 
                        group = nested_group_element(group, col_name, cell_value) #for the columns this works. 
            #you have this rows, thing creatd from the top.... 
            same_parent.append(group)

            #last_common = find_last_common_element(xml_group, group)


            
            else:
                #suddenly we are at parent_index 2... 
                #initial is still index 1, therefore xml group is still index 1
                last_common = find_last_common_element(xml_group, same_parent[0]) #same_parent arbitrarily chosen
                #this last common will be wm>>
                for gr in same_parent: 
                    xml_group = xml_group.find(".//" + last_common).append(gr)

                same_parent = []
                del xml_dict[initial_parent_index]
                xml_dict[]




                    last_common.append(gr)
                
                

                #append it


            

                    #the issue here is
                    #but for the NEXT ROW. we do not want to find the group sw2vc60, we want to create a new sw2
                    # 
                    # 
                    # so lets say for the next group_sw2vc60 row


                    #if sheet_xml_element == None:
                     #   sheet_xml_element = create_section(col_name, cell_value)
                    #else: 
                     #   sheet_xml_element = nested_group_element(sheet_xml_element, col_name, cell_value)

            #now you've created it sheet_xml_element for one row. 
            #you want to append this to the xml_group
            #you want to append S INTO W. 




            #next row, you will get relevent xml_group. 
            

                


            
            if parent_index != initial_parent_index:
                xml_dict[parent_index-1] 
                del xml_group

                #we have moved on... 
                #you are done working with this xml group. 
                #replace it in the xml_dict

            else:
                #

        
            
            parent_group = None
            if sheet_name != sheet_names[0]:
                parent_group = xml_dict.get(str(parent_index)) #initially this is going to be none, because there is nothing in the xml_dict, sheet 1
           
            if parent_group == None:
                print(sheet_name)
                print(parent_index)
            #if (sheet_name == sheet_names[1] and parent_group != None):
             #   first = ET.ElementTree(parent_group)
              #  first.write("help.xml") #first sheet, first xml, no S subgroup
            #next sheet
            #u will get a group based on the parent_index, pass that in and build upon it. 


            for col_name in question_headers: 
                col_num = headers.index(col_name)
                cell_value = str(row[col_num])

                if parent_group == None: #parent_group will be None, create from the TOPPPPPPPP. 
                    #parent_group will be None, for every row in the first sheet. 
                    parent_group = group_section(str(col_name), cell_value) #first sheet, first q going to be created



                else: #when parent group is not None, you have grabbed a parent. this needs to be second sheet onwards. 
                    parent_group = nested_group_element(parent_group, str(col_name), cell_value) #first sheet, second q, going to be assigned to group (THIS WHOLE ROW XML DONE)
        
            if sheet_name != sheet_names[0]:
                if parent_index in xml_dict:
                    del xml_dict[parent_index]
    
            #ok so this was correct
            #problem is that parent_index is the fuckin same for the first sheet. 


            #when you are in hte next sheet... 
            #thats when you need to start changing the values.

            #now you have this new, and edited parent group. 
            #put that in the dictionary with a new index. 

            xml_dict[index] = parent_group


                #first = ET.ElementTree(group)
                #first.write("first.xml") #first sheet, first xml, no S subgroup



            #index of the header, is the enumerated col_num-1

            #but for each row, we need to get the cell_value for the question_headers only

            print(xml_dict)
        #break #so if you bring this break statement back
        #if you only do ONE sheet, the elements under <wm> come back. but as soon as you do more. it gets overwritten. why.... \

            

            
                   for col_num, cell_value in enumerate(row, start=1):
                #only rows looked at will be for this submission
                
                col_name = str(headers[col_num - 1])
                if (col_name.startswith('_')):
                    continue


                #group_arr = col_name.split("/")
                if len(xml_dict) == 0: 
                    group = group_section(str(col_name), cell_value)
                    xml_dict[index] = group

                    
                    first = ET.ElementTree(group)
                    first.write("first.xml") #first sheet, first xml, no S subgroup
                     
                else:
                    print(xml_dict)
                    parent_group = xml_dict.get(parent_index)
                    if (parent_group == None):
                        group = group_section(str(col_name), cell_value)
                        xml_dict[index] = group
                    else:
                        group = nested_group_element(parent_group, str(col_name), cell_value)
                        del xml_dict[parent_index]
                        xml_dict[index] = group
      

    for group_xml in xml_dict.values():
        submission_xml.append(group_xml) 
    
    return submission_xml
        
    """


        #you want to check if there are multiple sheets 
        #go into the next sheet, pass in the index number in question. 
        #if an element has a parent_index of that value, thats what ur workin w (SO PARENT INDEX IS ALL 1 RN LETS SAY)

        #group_op5ep98/group_wm95d90/Enter_a_number_neste_does_not_repeat_tho

        # pass in boolean of new = True
        # pass in _uid lets say 
        #group op doesn't exist, so you create it
        #wm doesnt exist, so you create it, 
        #enter num doesnt exist, so you create it and add the cell value

        #new loop iter
        #pass in _uid, but this time, because its a new row 
        #do the same, create everything u just created

        #NEW SHEET

        #group_op5ep98/group_wm95d90/group_sw2vc60/nested_x_2_within_re_p_this_also_repeats
        #


        #ok how about this
        #create a list
        #you've been passed in an index already from the initial sheet 
        #look for the elements that are parent index, and create this section
        #put the op5 element in a list
        #so for the first row initial (first uuid), ur gonna have TWO elements in it (maybe have it be a dictionary or smth, cuz u want it to be easy lookup)



        #then the next sheet will be opend (next loop) 
        #loop through the vlaues in the 

        #now ur gonna go into the sheet, look for parent index 






"""
      #  _uid = ET.ElementTree(_uid)
      # _uid.write("uid.xml")

        # iterate through other sheets to create repeat groups, and append to xml
        repeat_elements = repeat_groups(_uid, formatted_uuid, workbook)
#        _repeat = ET.ElementTree(repeat_elements)
 #       _repeat.write("xm.xml")

        if repeat_elements != None:
            _uid = repeat_elements

        version = ET.Element("__version__")
        version.text = __version__
        _uid.append(version)

        formatted_uuid = meta_element(_uid, formatted_uuid)

        # for initial transfer (without uuids), attachments are saved with row numbers
        # row number folder is renamed to uuid to complete transfer to kobo and associate attachment to specific response
        rename_media_folder(submission_data, formatted_uuid[len("uuid:") :], row_num)
        results.append(_uid)
        num_results += 1


    count = ET.SubElement(root, "count")
    count.text = str(num_results)
    next = ET.SubElement(root, "next")
    next.text = None
    previous = ET.SubElement(root, "previous")
    previous.text = None
    root.append(results)

    root = ET.ElementTree(root)
    root.write("./um.xml")
    workbook.close()
    
    return root

"""