# Added by Yu Tsukioka 17OCT2024 for downloading KoBoToolbox submissions in various formats

import requests
import json
import sys
import argparse
from pathlib import Path
import pandas as pd
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import time

current_dir = Path(__file__).resolve().parent
project_root = current_dir.parent
sys.path.insert(0, str(project_root))
from helpers.config import Config

def get_form_id(config):
    try:
        response = requests.get(
            f"{config['kc_url']}/api/v1/forms",
            headers=config['headers']
        )
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"❌ Error retrieving form ID: {e}")
        sys.exit(1)

    forms = response.json()
    for form in forms:
        if form['id_string'] == config['asset_uid']:
            return form['formid']
    print("❌ Form ID not found.")
    sys.exit(1)

def download_xlsx(asset_uid, form_id, kf_url, headers, output_dir):
    xlsx_url = f"{kf_url}/api/v1/data/{form_id}.xlsx"
    print(f"📥 Downloading XLSX file from: {xlsx_url}")

    try:
        response = requests.get(xlsx_url, headers=headers, stream=True)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"❌ Error downloading XLSX file: {e}")
        sys.exit(1)

    output_file = output_dir / f"{asset_uid}.xlsx"

    with open(output_file, 'wb') as f:
        for chunk in response.iter_content(chunk_size=1024):
            if chunk:
                f.write(chunk)

    print(f"✅ XLSX file saved successfully to: {output_file}")
    return output_file

def download_xlsx_text(asset_uid, kf_url, headers, output_dir):
    """
    Download XLSX with text formatting using the export API.
    """
    create_payload = {
        "type": "xlsx",
        "xls_types_as_text": True,
        "include_labels": False,
        "xpath": "results"
    }
    create_url = f"{kf_url}/api/v2/assets/{asset_uid}/exports/"
    print(f"📥 Creating XLSX export with text formatting...")
    
    try:
        r = requests.post(create_url, headers=headers, json=create_payload)
        r.raise_for_status()
        export_info = r.json()
    except requests.exceptions.RequestException as e:
        print(f"❌ Error creating export: {e}")
        sys.exit(1)

    export_url = export_info["url"]
    print(f"📥 Waiting for export to complete...")
    while export_info["state"] != "complete":
        time.sleep(3)
        try:
            export_info = requests.get(export_url, headers=headers).json()
            if export_info["state"] == "error":
                print(f"❌ Export failed: {export_info.get('messages', 'Unknown error')}")
                sys.exit(1)
        except requests.exceptions.RequestException as e:
            print(f"❌ Error checking export status: {e}")
            sys.exit(1)

    # 3. download the final file
    dl_url = export_info["result"]["url"]
    out_file = output_dir / f"{asset_uid}_text.xlsx"
    
    print(f"📥 Downloading XLSX file with text formatting...")
    try:
        with requests.get(dl_url, headers=headers, stream=True) as resp:
            resp.raise_for_status()
            with open(out_file, "wb") as f:
                for chunk in resp.iter_content(1024):
                    if chunk:
                        f.write(chunk)
    except requests.exceptions.RequestException as e:
        print(f"❌ Error downloading XLSX file: {e}")
        sys.exit(1)
    
    print(f"✅ XLSX file with text formatting saved to: {out_file}")
    return out_file

def download_xml(asset_uid, kf_url, headers, output_dir):
    download_url = f"{kf_url}/api/v2/assets/{asset_uid}/data.xml"
    print(f"📥 Downloading XML data from: {download_url}")

    try:
        response = requests.get(download_url, headers=headers, stream=True)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"❌ Error downloading XML file: {e}")
        sys.exit(1)

    output_file = output_dir / f"{asset_uid}.xml"

    with open(output_file, 'wb') as f:
        for chunk in response.iter_content(chunk_size=1024):
            if chunk:
                f.write(chunk)

    print(f"✅ XML file saved successfully to: {output_file}")
    return output_file

def download_other_format(asset_uid, kf_url, headers, data_format, output_dir):
    download_url = f"{kf_url}/api/v2/assets/{asset_uid}/data.{data_format}"
    print(f"📥 Downloading {data_format.upper()} data from: {download_url}")

    try:
        response = requests.get(download_url, headers=headers)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"❌ Error downloading {data_format} file: {e}")
        sys.exit(1)

    output_file = output_dir / f"{asset_uid}.{data_format}"

    if data_format.lower() == "json":
        try:
            data = response.json()
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4, ensure_ascii=False)
            print(f"✅ {data_format.upper()} file saved successfully to: {output_file}")
            return output_file
        except ValueError as e:
            print(f"❌ Error processing JSON data: {e}")
            sys.exit(1)
    else:
        with open(output_file, 'wb') as f:
            for chunk in response.iter_content(chunk_size=1024):
                if chunk:
                    f.write(chunk)
        print(f"✅ {data_format.upper()} file saved successfully to: {output_file}")
        return output_file
        
def generate_validation_status(output_file, output_dir, output_json):
    try:
        df_xlsx = pd.read_excel(output_file)

        if 'meta/instanceID' not in df_xlsx.columns:
            print("❌ 'meta/instanceID' column not found in XLSX.")
            sys.exit(1)

        df_xlsx['_uuid'] = df_xlsx['meta/instanceID'].str.replace('uuid:', '', regex=False)

        with open(output_json, 'r', encoding='utf-8') as f:
            data_json = json.load(f)

        records = []
        for entry in data_json.get("results", []):
            meta_instance_id = entry.get("meta/instanceID", "")
            if meta_instance_id.startswith("uuid:"):
                uuid = meta_instance_id.split("uuid:")[1]
            else:
                uuid = meta_instance_id
            record = {
                "_uuid": uuid,
                "_validation_status": entry.get("_validation_status", {}).get("uid"),
                "_id": entry.get("_id"),
            }
            records.append(record)

        df_json = pd.DataFrame(records)

        df_xlsx["_uuid"] = df_xlsx["_uuid"].astype(str)
        df_json["_uuid"] = df_json["_uuid"].astype(str)

        merged_df = pd.merge(df_xlsx, df_json, on="_uuid", how="left", suffixes=('', '_json'))

        if merged_df['_validation_status'].isnull().all():
            print("❌ No matching '_uuid' found between XLSX and JSON data.")
            sys.exit(1)

        relevant_columns = ['_uuid', '_validation_status', '_id_json']
        if not all(col in merged_df.columns for col in relevant_columns):
            print(f"❌ One or more required columns {relevant_columns} are missing in the merged DataFrame.")
            sys.exit(1)

        validation_df = merged_df[relevant_columns].rename(columns={'_validation_status': 'validation_status_uid'})

        wb = Workbook()
        ws = wb.active
        ws.title = "Validation Status"
        ws.append(["_uuid", "validation_status_uid", "_id"])

        for index, row in validation_df.iterrows():
            ws.append([row['_uuid'], row['validation_status_uid'], row['_id_json']])

        validation_options = ["validation_status_approved", "validation_status_not_approved", "validation_status_on_hold"]

        validation_rule = DataValidation(
            type="list",
            formula1=f'"{",".join(validation_options)}"',
            allow_blank=True,
            showDropDown=True
        )

        for row_num in range(2, ws.max_row + 1):
            cell = f"{get_column_letter(2)}{row_num}"
            validation_rule.add(ws[cell])

        ws.add_data_validation(validation_rule)

        val_stat_output_file = output_dir / f"{Path(output_file).stem}_val_stat.xlsx"
        wb.save(val_stat_output_file)
        print(f"✅ Generated validation status Excel: {val_stat_output_file}")
    except Exception as e:
        print(f"❌ Error generating validation status Excel: {e}")
        sys.exit(1)
                                              
def main(config_file=None, format=None, text_format=False, skip_media_download=False):
    try:
        config = Config(config_file=config_file).src
    except Exception as e:
        print(f"⚠️ Failed to load configuration: {e}")
        sys.exit(1)

    asset_uid = config['asset_uid']
    headers = config['headers']
    kf_url = config['kf_url']

    output_dir = Path("attachments")
    output_dir.mkdir(exist_ok=True)

    if format == "xlsx":
        if text_format:
            output_xlsx = download_xlsx_text(asset_uid, kf_url, headers, output_dir)
        else:
            form_id = get_form_id(config)
            output_xlsx = download_xlsx(asset_uid, form_id, kf_url, headers, output_dir)
        
        output_json = download_other_format(asset_uid, kf_url, headers, "json", output_dir)
        generate_validation_status(output_xlsx, output_dir, output_json)
    else:
        download_other_format(asset_uid, kf_url, headers, format, output_dir)
    
    if skip_media_download:
        print("📋 Skipping media file downloads as requested")

if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='A CLI tool to download data from KoBoToolbox in various formats.'
    )
    parser.add_argument(
        "--config-file",
        "-c",
        type=str,
        default="config.json",
        help="Path to the configuration file (default: config.json)"
    )
    parser.add_argument(
        "--format",
        "-f",
        type=str,
        choices=["xml", "geojson", "json", "xlsx"],
        required=True,
        help="Data format: 'xml', 'geojson', 'json', or 'xlsx'."
    )
    parser.add_argument(
        "--text-format",
        "-t",
        default=False,
        action='store_true',
        help="For XLSX format, preserve all data as text (avoid automatic type conversion)."
    )
    parser.add_argument(
        "--skip-media-download",
        "-smd",
        default=False,
        action='store_true',
        help="Skip downloading media files, only download submission data."
    )
    args = parser.parse_args()

    try:
        main(
            config_file=args.config_file,
            format=args.format,
            text_format=args.text_format,
            skip_media_download=args.skip_media_download,
        )
    except KeyboardInterrupt:
        print('🛑 Stopping run')
        sys.exit()