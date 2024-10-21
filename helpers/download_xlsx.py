import requests
import os
import sys
import argparse
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.abspath(os.path.join(current_dir, '..'))
sys.path.insert(0, project_root)
from helpers.config import Config
import pandas as pd
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def get_form_id(config):
    try:
        response = requests.get(
            f"{config['kc_url']}/api/v1/forms",
            headers=config['headers']
        )
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"❌ Error retrieving form ID: {e}")
        sys.exit(1)

    forms = response.json()
    for form in forms:
        if form['id_string'] == config['asset_uid']:
            return form['formid']
    print("❌ Form ID not found.")
    sys.exit(1)

def download_xlsx(config_file):
    try:
        config = Config(config_file=config_file).src
    except Exception as e:
        print(f"⚠️ Failed to load configuration: {e}")
        sys.exit(1)
    asset_uid = config['asset_uid']
    kc_url = config['kf_url']
    headers = config['headers']
    form_id = get_form_id(config)
    xlsx_url = f"{kc_url}/api/v1/data/{form_id}.xlsx"
    
    print(f"📥 Downloading XLSX file from: {xlsx_url}")
    try:
        response = requests.get(xlsx_url, headers=headers, stream=True)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"❌ Error downloading file: {e}")
        sys.exit(1)

    output_dir = "attachments"
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f"{asset_uid}.xlsx")

    with open(output_file, 'wb') as f:
        for chunk in response.iter_content(chunk_size=1024):
            if chunk:
                f.write(chunk)

    print(f"✅ XLSX file saved successfully to: {output_file} of FormID: {form_id}")
    try:
        df = pd.read_excel(output_file)
        new_df = pd.DataFrame({
            '_uuid': df['_uuid'],
            'validation_status_uid': ''
        })
        wb = Workbook()
        ws = wb.active
        ws.title = "Validation Status"
        ws.append(["_uuid", "validation_status_uid"])
        for row in new_df.itertuples(index=False, name=None):
            uuid_value = row[0]
            ws.append([uuid_value, ""])

        validation_options = ["", "validation_status_approved", "validation_status_not_approved", "validation_status_on_hold"]

        validation_rule = DataValidation(
            type="list",
            formula1=f'"{",".join(validation_options)}"',
            allow_blank=True,
            showDropDown=True
        )
        
        validation_column = get_column_letter(2)
        for row in range(2, ws.max_row + 1):
            ws[f"{validation_column}{row}"] = ""
            validation_rule.add(ws[f"{validation_column}{row}"])
        
        ws.add_data_validation(validation_rule)
        
        val_stat_output_file = os.path.join("attachments", f"{os.path.splitext(os.path.basename(output_file))[0]}_val_stat.xlsx")
        wb.save(val_stat_output_file)

        print(f"✅ Generated validation status Excel: {val_stat_output_file}")
    except Exception as e:
        print(f"❌ Error generating validation status Excel: {e}")
        sys.exit(1)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Download XLSX export from KoBoToolbox"
    )
    parser.add_argument(
        "--config-file",
        "-c",
        type=str, 
        default="config.json", 
        help="Path to the configuration file (default: config.json)"
    )
    args = parser.parse_args()

    download_xlsx(args.config_file)
